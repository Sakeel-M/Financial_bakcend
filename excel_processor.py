"""
Excel processor for UAE bank statements without pandas dependency
Uses openpyxl directly for better Excel compatibility
"""
import json
import re
from datetime import datetime
import os
import openai

try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Set OpenAI API key
openai.api_key = os.getenv('OPENAI_API_KEY')

class UAEBankExcelProcessor:
    def __init__(self):
        self.bank_patterns = {
            'ADCB': ['abu dhabi commercial bank', 'adcb'],
            'FAB': ['first abu dhabi bank', 'fab'],
            'ENBD': ['emirates nbd', 'enbd'],
            'MASHREQ': ['mashreq bank', 'mashreq'],
            'CBD': ['commercial bank of dubai', 'cbd'],
            'HSBC': ['hsbc'],
            'RAKBANK': ['rak bank', 'rakbank'],
            'ADIB': ['abu dhabi islamic bank', 'adib'],
            'BOA': ['bank of america', 'bofa', 'boa', 'b of a'],
            'CHASE': ['chase', 'jp morgan chase', 'jpmorgan'],
            'WELLS': ['wells fargo', 'wells'],
            'CITI': ['citibank', 'citi']
        }

        self.categories = {
            'Food & Dining': [
                'carrefour', 'lulu', 'spinneys', 'choithrams', 'union coop', 'waitrose',
                'restaurant', 'cafe', 'kfc', 'mcdonald', 'pizza', 'subway', 'dominos',
                'starbucks', 'costa', 'dunkin', 'burger', 'food', 'dining', 'eat',
                'grocery', 'supermarket', 'hypermarket', 'bakery', 'deli', 'bistro',
                'catering', 'takeaway', 'delivery', 'zomato', 'talabat', 'deliveroo',
                'doordash', 'grubhub', 'uber eats', 'postmates', 'chipotle', 'panera',
                'whole foods', 'trader joe', 'safeway', 'kroger', 'target', 'walmart',
                'donuts', 'coffee', 'espresso', 'bacio di latte', 'butchery', 'zinque',
                'ralphs', 'albertsons', 'vons', 'pavilions', 'publix', 'wegmans', 'harris teeter'
            ],
            'Transportation': [
                'adnoc', 'eppco', 'enoc', 'petrol', 'fuel', 'gas', 'gasoline',
                'taxi', 'uber', 'careem', 'metro', 'bus', 'rta', 'parking',
                'salik', 'toll', 'car wash', 'transport', 'emirates', 'etihad',
                'flydubai', 'air arabia', 'airline', 'flight', 'airport',
                'chevron', 'shell', 'bp', '76', 'exxon', 'mobil', 'lyft',
                'parking services', 'toll roads', 'valet'
            ],
            'Shopping & Retail': [
                'mall', 'centrepoint', 'max', 'home centre', 'ikea', 'ace',
                'sharaf dg', 'jumbo', 'electronics', 'clothing', 'fashion',
                'shop', 'store', 'retail', 'amazon', 'noon', 'souq', 'namshi',
                'h&m', 'zara', 'nike', 'adidas', 'apple', 'samsung', 'virgin',
                'uniqlo', 'target', 'walmart', 'costco', 'best buy', 'macy',
                'nordstrom', 'kohls', 'tj maxx', 'ross', 'marshalls'
            ],
            'Healthcare': [
                'hospital', 'clinic', 'pharmacy', 'medical', 'doctor', 'health',
                'dental', 'medicare', 'aster', 'nmc', 'mediclinic', 'life pharmacy',
                'boots', 'aster pharmacy', 'day today pharmacy'
            ],
            'Utilities & Bills': [
                'dewa', 'addc', 'sewa', 'fewa', 'etisalat', 'du', 'internet',
                'mobile', 'telecom', 'electricity', 'water', 'gas', 'utility', 'bill',
                'wifi', 'broadband', 'phone bill', 'electric bill'
            ],
            'Entertainment': [
                'cinema', 'movie', 'vox', 'reel', 'netflix', 'osn', 'gaming',
                'entertainment', 'park', 'beach', 'attraction', 'ticket', 'event',
                'spotify', 'youtube', 'disney', 'amazon prime', 'hulu', 'shahid',
                'disneyland', 'balloon museum', 'sawdust festival', 'museum', 'amusement'
            ],
            'Subscriptions & Digital Services': [
                'netflix', 'spotify', 'youtube premium', 'amazon prime', 'disney+',
                'adobe', 'microsoft', 'google', 'icloud', 'dropbox', 'zoom',
                'subscription', 'monthly', 'annual', 'recurring', 'saas',
                'software', 'app store', 'play store', 'itunes', 'office 365'
            ],
            'ATM & Cash Withdrawals': [
                'atm', 'cash withdrawal', 'withdrawal', 'atm withdrawal',
                'cash advance', 'atm fee', 'withdrawal fee'
            ],
            'Banking & Finance': [
                'transfer', 'fee', 'charge', 'finance', 'loan', 'interest',
                'bank fee', 'service charge', 'maintenance fee', 'overdraft',
                'wire transfer', 'remittance', 'exchange'
            ],
            'Personal Care': [
                'salon', 'spa', 'barbershop', 'beauty', 'cosmetics', 'skincare',
                'haircut', 'manicure', 'pedicure', 'massage', 'gym', 'fitness',
                '24hourfitness', 'planet fitness', 'la fitness', 'crunch', 'equinox',
                'flex fitness', 'orangetheory'
            ]
        }

    def detect_bank(self, text):
        """Detect bank from text content"""
        text_lower = text.lower()
        for bank_code, patterns in self.bank_patterns.items():
            for pattern in patterns:
                if pattern in text_lower:
                    return {
                        'ADCB': 'Abu Dhabi Commercial Bank',
                        'FAB': 'First Abu Dhabi Bank',
                        'ENBD': 'Emirates NBD',
                        'MASHREQ': 'Mashreq Bank',
                        'CBD': 'Commercial Bank of Dubai',
                        'HSBC': 'HSBC Bank Middle East',
                        'RAKBANK': 'RAKBank',
                        'ADIB': 'Abu Dhabi Islamic Bank',
                        'BOA': 'Bank of America',
                        'CHASE': 'Chase Bank',
                        'WELLS': 'Wells Fargo',
                        'CITI': 'Citibank'
                    }.get(bank_code, f'{bank_code} Bank')
        return 'Unknown Bank'

    def categorize_transaction(self, description):
        """Enhanced categorization with priority matching and better logic"""
        desc_lower = description.lower().strip()

        # Priority-based categorization for better accuracy
        category_priority = [
            'ATM & Cash Withdrawals',
            'Subscriptions & Digital Services',
            'Food & Dining',
            'Transportation',
            'Healthcare',
            'Utilities & Bills',
            'Entertainment',
            'Shopping & Retail',
            'Personal Care',
            'Banking & Finance'
        ]

        # Check for exact matches first
        for category in category_priority:
            keywords = self.categories[category]
            for keyword in keywords:
                if keyword in desc_lower:
                    # Additional validation for specific categories
                    if category == 'ATM & Cash Withdrawals':
                        if any(term in desc_lower for term in ['atm', 'withdrawal', 'cash']):
                            return category
                    elif category == 'Subscriptions & Digital Services':
                        if any(term in desc_lower for term in ['subscription', 'monthly', 'netflix', 'spotify', 'prime', 'office', 'adobe', 'google', 'microsoft', 'icloud']):
                            return category
                    elif category == 'Banking & Finance':
                        # Only categorize as banking if it's clearly a bank transaction
                        if any(term in desc_lower for term in ['transfer', 'fee', 'charge', 'interest', 'maintenance']):
                            return category
                    else:
                        return category

        # Additional pattern matching for common transaction types
        if any(word in desc_lower for word in ['taxi', 'uber', 'careem', 'rta']):
            return 'Transportation'

        if any(word in desc_lower for word in ['restaurant', 'cafe', 'food', 'dining', 'eat']):
            return 'Food & Dining'

        if any(word in desc_lower for word in ['mall', 'shop', 'store', 'retail']):
            return 'Shopping & Retail'

        # Check for merchant names and common patterns
        merchant_patterns = {
            'CAREEM': 'Transportation',
            'TALABAT': 'Food & Dining',
            'ZOMATO': 'Food & Dining',
            'NETFLIX': 'Subscriptions & Digital Services',
            'SPOTIFY': 'Subscriptions & Digital Services',
            'AMAZON': 'Shopping & Retail',
            'NOON': 'Shopping & Retail',
            'ADNOC': 'Transportation',
            'ENOC': 'Transportation',
            'EPPCO': 'Transportation'
        }

        for merchant, cat in merchant_patterns.items():
            if merchant.lower() in desc_lower:
                return cat

        return 'Other Expenses'

    def ai_categorize_transactions(self, transactions):
        """Use AI to intelligently categorize transactions in batches"""
        if not openai.api_key or not transactions:
            return transactions

        # Define available categories
        categories = [
            'Food & Dining',
            'Transportation',
            'Shopping & Retail',
            'Healthcare',
            'Utilities & Bills',
            'Entertainment',
            'Subscriptions & Digital Services',
            'ATM & Cash Withdrawals',
            'Banking & Finance',
            'Personal Care',
            'Travel',
            'Income',
            'Other Expenses'
        ]

        try:
            # Process in batches of 50 transactions
            batch_size = 50
            categorized_transactions = []

            for i in range(0, len(transactions), batch_size):
                batch = transactions[i:i + batch_size]

                # Prepare batch for AI
                descriptions = [f"{idx}. {tx['Description']}" for idx, tx in enumerate(batch, start=i)]

                prompt = f"""Categorize these financial transactions into the most appropriate category. READ THE DESCRIPTION CAREFULLY and categorize accurately.

Available categories: {', '.join(categories)}

STRICT Category Rules - Follow These Examples:

1. Food & Dining - RESTAURANTS, CAFES, GROCERY STORES:
   ✓ CAVA, Chipotle, McDonald's, Panera, Five Guys, In-N-Out
   ✓ Starbucks, Coffee Bean, Dutch Bros, any coffee shop
   ✓ Ralphs, Kroger, Safeway, Albertsons, Vons, Whole Foods, Trader Joe's, Target (food section), Walmart (food section)
   ✓ DoorDash, UberEats, Grubhub, Postmates
   ✓ Bakeries, donut shops, butcher shops, delis
   ✓ SQ *BACIO DI LATTE, SQ *MARU ESPRESSO BAR, DD *DOORDASH

2. Transportation - GAS, PARKING, RIDESHARE:
   ✓ Chevron, Shell, Mobil, BP, 76, Exxon, Texaco
   ✓ Uber (rides), Lyft, taxi services
   ✓ LAX parking, valet parking, parking services, parking meters
   ✓ Toll roads, FasTrak, E-ZPass
   ✓ Car wash, auto services
   ✗ NOT food delivery apps

3. Utilities & Bills - PHONE, INTERNET, ELECTRICITY:
   ✓ Verizon, AT&T, T-Mobile, Sprint
   ✓ Comcast, Spectrum, Cox Internet
   ✓ Electric company, water company, gas utility
   ✗ NOT Rocket Money (that's Subscriptions & Digital Services)

4. Subscriptions & Digital Services - MONTHLY RECURRING SERVICES:
   ✓ Netflix, Hulu, Disney+, HBO Max, Amazon Prime Video
   ✓ Spotify, YouTube Premium, Apple Music
   ✓ Rocket Money, Mint, YNAB (budgeting apps)
   ✓ Adobe Creative Cloud, Microsoft Office 365, Dropbox
   ✓ Amazon Prime membership, Costco membership
   ✓ Any app subscription, cloud storage, SaaS

5. Personal Care - GYM, SALON, SPA:
   ✓ 24HourFitness, Planet Fitness, LA Fitness, Gold's Gym, Flex Fitness
   ✓ Hair salons, barbershops, nail salons
   ✓ Spa services, massage therapy
   ✓ Beauty products, cosmetics (if from beauty store)

6. Entertainment - MOVIES, EVENTS, ATTRACTIONS:
   ✓ Movie theaters (AMC, Regal), concerts, sports events
   ✓ Theme parks (Disneyland, Six Flags), museums (Balloon Museum)
   ✓ Festivals (Sawdust Festival), entertainment venues

7. Banking & Finance - BANK FEES, TRANSFERS:
   ✓ Zelle payments, wire transfers, bank transfers
   ✓ Bank fees, service charges, overdraft fees
   ✓ Foreign transaction fees, ATM fees (if separate from withdrawal)
   ✗ NOT ATM withdrawals (that's ATM & Cash Withdrawals)

8. Shopping & Retail - CLOTHING, ELECTRONICS, GENERAL MERCHANDISE:
   ✓ Uniqlo, H&M, Zara, Nike, Adidas (clothing)
   ✓ Target (non-food), Walmart (non-food), Amazon (general)
   ✓ Best Buy, Apple Store (electronics)
   ✗ NOT grocery stores

9. ATM & Cash Withdrawals:
   ✓ ATM Withdrawal, Cash Advance, ATM cash
   ✗ Nothing else

10. Income:
   ✓ Salary, paycheck deposits, direct deposits
   ✓ Refunds, reimbursements, tax refunds
   ✓ Money received (NOT sent)

11. Travel:
   ✓ Hotels, Airbnb, Booking.com
   ✓ Airlines (United, Delta, Southwest)
   ✓ Rental cars (Hertz, Enterprise)

12. Healthcare:
   ✓ Hospitals, clinics, doctor visits
   ✓ Pharmacies (CVS, Walgreens for prescriptions)
   ✓ Dental, vision care

13. Other Expenses:
   ✓ ONLY use this if truly unidentifiable

CRITICAL EXAMPLES TO LEARN FROM:
- "Rocket Money DES" → Subscriptions & Digital Services (budgeting app, NOT utilities)
- "SQ *BACIO DI LATTE" → Food & Dining (coffee shop)
- "CHECKCARD A PARKING SERVICES" → Transportation (parking)
- "24HourFitness" → Personal Care (gym)
- "Zelle payment to Bryan" → Banking & Finance (transfer)
- "TARGET ST 4255" → Check description - food items = Food & Dining, else Shopping & Retail
- "RALPHS #0299" → Food & Dining (grocery store)
- "Netflix" → Subscriptions & Digital Services
- "Verizon Wireless" → Utilities & Bills (phone service)

Transactions to categorize:
{chr(10).join(descriptions[:30])}

Respond ONLY with a JSON array of category names in the EXACT order of transactions, nothing else.
Example format: ["Food & Dining", "Transportation", "Shopping & Retail"]"""

                try:
                    response = openai.chat.completions.create(
                        model="gpt-4o-mini",
                        messages=[
                            {
                                "role": "system",
                                "content": "You are a financial categorization expert. Respond only with a JSON array of categories, no additional text."
                            },
                            {"role": "user", "content": prompt}
                        ],
                        max_tokens=500,
                        temperature=0.1
                    )

                    ai_response = response.choices[0].message.content.strip()

                    # Remove markdown if present
                    if ai_response.startswith('```json'):
                        ai_response = ai_response.replace('```json', '').replace('```', '').strip()
                    elif ai_response.startswith('```'):
                        ai_response = ai_response.replace('```', '').strip()

                    ai_categories = json.loads(ai_response)

                    # Apply AI categories to batch
                    for idx, tx in enumerate(batch):
                        if idx < len(ai_categories):
                            tx['Category'] = ai_categories[idx]
                            tx['Subcategory'] = ai_categories[idx].split(' ')[0]
                        categorized_transactions.append(tx)

                except Exception as e:
                    print(f"AI categorization error for batch: {str(e)}")
                    # Fallback to rule-based for this batch
                    for tx in batch:
                        categorized_transactions.append(tx)

            return categorized_transactions

        except Exception as e:
            print(f"AI categorization failed: {str(e)}")
            return transactions

    def extract_bank_info(self, worksheet):
        """Extract bank information from worksheet"""
        bank_info = {
            'bank_name': 'UAE Bank',
            'account_holder': 'Account Holder',
            'account_number': 'XXXX-XXXX-XXXX',
            'currency': 'AED'
        }

        # Check first 20 rows for bank information
        for row in range(1, min(21, worksheet.max_row + 1)):
            for col in range(1, min(10, worksheet.max_column + 1)):
                cell_value = str(worksheet.cell(row, col).value or '').upper()

                # Detect bank
                bank_name = self.detect_bank(cell_value)
                if bank_name not in ['UAE Bank', 'Unknown Bank']:
                    bank_info['bank_name'] = bank_name

                # Look for account holder (names are usually 2-4 words, mixed case)
                if len(cell_value.split()) >= 2 and len(cell_value.split()) <= 4:
                    if not any(char.isdigit() for char in cell_value):
                        if any(word in cell_value.lower() for word in ['mr', 'ms', 'mrs', 'dr']):
                            bank_info['account_holder'] = cell_value.title()

                # Look for account numbers
                if re.match(r'^\d{10,16}$', cell_value.replace(' ', '').replace('-', '')):
                    masked = cell_value[:4] + '-' + '*' * 4 + '-' + cell_value[-4:]
                    bank_info['account_number'] = masked

        return bank_info

    def find_data_headers(self, worksheet):
        """Find the row with data headers and detect column mappings"""
        header_keywords = ['date', 'amount', 'description', 'particular', 'narration', 'debit', 'credit', 'balance', 'type', 'reference']

        for row in range(1, min(20, worksheet.max_row + 1)):
            row_data = {}
            header_count = 0

            for col in range(1, min(worksheet.max_column + 1, 10)):  # Check up to 10 columns
                cell_value = str(worksheet.cell(row, col).value or '').lower().strip()

                # Map common header variations
                if cell_value and any(keyword in cell_value for keyword in header_keywords):
                    header_count += 1

                    # Detect column types
                    if 'date' in cell_value:
                        row_data['date_col'] = col
                    elif 'description' in cell_value or 'particular' in cell_value or 'narration' in cell_value:
                        row_data['description_col'] = col
                    elif 'type' in cell_value:
                        row_data['type_col'] = col
                    elif 'reference' in cell_value or 'ref' in cell_value:
                        row_data['reference_col'] = col
                    elif 'debit' in cell_value:
                        row_data['debit_col'] = col
                    elif 'credit' in cell_value:
                        row_data['credit_col'] = col
                    elif 'amount' in cell_value:
                        row_data['amount_col'] = col

            # If we found at least 3 header columns, this is likely the header row
            if header_count >= 3:
                row_data['header_row'] = row
                return row_data

        # Default mapping if no headers found
        return {'header_row': 1, 'date_col': 1, 'description_col': 2, 'debit_col': 4, 'credit_col': 5}

    def detect_date_format(self, date_string):
        """Detect and parse date from various formats"""
        if isinstance(date_string, datetime):
            return date_string.strftime('%Y-%m-%d')

        date_str = str(date_string).strip()

        # Try multiple date format patterns
        patterns = [
            # DD/MM/YYYY (UAE format)
            (r'^(\d{1,2})/(\d{1,2})/(\d{4})$', lambda m: f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"),
            # MM/DD/YY (US format - 2 digit year)
            (r'^(\d{1,2})/(\d{1,2})/(\d{2})$', lambda m: f"20{m.group(3)}" if int(m.group(3)) < 50 else f"19{m.group(3)}" + f"-{m.group(1).zfill(2)}-{m.group(2).zfill(2)}"),
            # MM/DD/YYYY (US format - 4 digit year)
            (r'^(\d{1,2})/(\d{1,2})/(\d{4})$', lambda m: f"{m.group(3)}-{m.group(1).zfill(2)}-{m.group(2).zfill(2)}"),
            # YYYY-MM-DD (ISO format)
            (r'^(\d{4})-(\d{1,2})-(\d{1,2})$', lambda m: f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"),
        ]

        for pattern, formatter in patterns:
            match = re.match(pattern, date_str)
            if match:
                try:
                    # Special handling for MM/DD/YY format
                    if pattern == r'^(\d{1,2})/(\d{1,2})/(\d{2})$':
                        month, day, year = match.group(1), match.group(2), match.group(3)
                        full_year = '20' + year if int(year) < 50 else '19' + year
                        return f"{full_year}-{month.zfill(2)}-{day.zfill(2)}"
                    else:
                        return formatter(match)
                except:
                    continue

        # Fallback to current year if parsing fails
        return datetime.now().strftime('%Y-%m-%d')

    def process_excel_file(self, file_content):
        """Dynamic Excel processor - automatically detects and adapts to different Excel formats"""
        if not EXCEL_AVAILABLE:
            return None, "openpyxl not available - using demo data"

        try:
            # Load workbook
            workbook = load_workbook(file_content)
            print(f"[*] Processing Excel file with {len(workbook.sheetnames)} sheets")

            # Extract bank info from Account Info sheet
            bank_info = {'bank_name': 'Unknown Bank', 'account_holder': 'Account Holder', 'account_number': 'XXXX-XXXX-XXXX', 'currency': 'USD'}
            if 'Account Info' in workbook.sheetnames:
                info_sheet = workbook['Account Info']
                for row in range(1, info_sheet.max_row + 1):
                    field = str(info_sheet.cell(row, 1).value or '').lower()
                    value = str(info_sheet.cell(row, 2).value or '')
                    if 'account holder' in field and value:
                        bank_info['account_holder'] = value
                    elif 'account number' in field and value:
                        bank_info['account_number'] = value
                    elif 'bank name' in field and value:
                        detected_bank = self.detect_bank(value)
                        if detected_bank != 'Unknown Bank':
                            bank_info['bank_name'] = detected_bank
                            # Set currency based on bank
                            if 'America' in detected_bank or 'Chase' in detected_bank or 'Wells' in detected_bank or 'Citi' in detected_bank:
                                bank_info['currency'] = 'USD'
                            else:
                                bank_info['currency'] = 'AED'

            print(f"[BANK] {bank_info['bank_name']} | Currency: {bank_info['currency']}")

            # Process transaction sheets (all sheets except Account Info)
            all_transactions = []
            transaction_sheets = [sheet for sheet in workbook.sheetnames if sheet != 'Account Info']

            for sheet_name in transaction_sheets:
                worksheet = workbook[sheet_name]
                print(f"\n[SHEET] Processing: {sheet_name}")

                # DYNAMIC HEADER DETECTION - automatically find headers and column mappings
                column_map = self.find_data_headers(worksheet)
                header_row = column_map.get('header_row', 1)
                date_col = column_map.get('date_col', 1)
                desc_col = column_map.get('description_col', 2)
                debit_col = column_map.get('debit_col', 4)
                credit_col = column_map.get('credit_col', 5)
                amount_col = column_map.get('amount_col', None)

                print(f"  [OK] Headers found at row {header_row}")
                print(f"  [OK] Columns: Date={date_col}, Desc={desc_col}, Debit={debit_col}, Credit={credit_col}")

                # Process transactions starting from row after headers
                sheet_transactions = 0
                for row in range(header_row + 1, worksheet.max_row + 1):
                    try:
                        # DYNAMIC CELL READING - read based on detected column positions
                        date_cell = worksheet.cell(row, date_col).value if date_col else None
                        description_cell = worksheet.cell(row, desc_col).value if desc_col else None
                        debit_cell = worksheet.cell(row, debit_col).value if debit_col else None
                        credit_cell = worksheet.cell(row, credit_col).value if credit_col else None

                        # Skip empty rows or summary rows
                        if not date_cell:
                            continue
                        if any(keyword in str(date_cell).upper() for keyword in ['TOTAL', 'SUMMARY', 'BALANCE', 'OPENING', 'CLOSING']):
                            continue
                        if description_cell and any(keyword in str(description_cell).upper() for keyword in ['TOTAL', 'SUMMARY', 'MONTHLY']):
                            continue

                        # INTELLIGENT DATE PARSING - automatically detect and parse date format
                        date_str = self.detect_date_format(date_cell)

                        # DYNAMIC AMOUNT EXTRACTION - handle both debit/credit and single amount columns
                        amount = 0
                        try:
                            if amount_col:
                                # Single amount column
                                amount = float(str(worksheet.cell(row, amount_col).value or '0').replace(',', ''))
                            else:
                                # Debit/Credit columns
                                if debit_cell and str(debit_cell).strip() and str(debit_cell).strip() not in ['', '-', 'None']:
                                    amount = -abs(float(str(debit_cell).replace(',', '')))  # Debits are negative
                                elif credit_cell and str(credit_cell).strip() and str(credit_cell).strip() not in ['', '-', 'None']:
                                    amount = abs(float(str(credit_cell).replace(',', '')))  # Credits are positive
                        except (ValueError, TypeError):
                            continue

                        # Skip if no amount
                        if amount == 0:
                            continue

                        # Process description
                        description = str(description_cell).strip() if description_cell else f"Transaction {len(all_transactions) + 1}"

                        # Categorize transaction (will be improved by AI later)
                        category = self.categorize_transaction(description)

                        transaction = {
                            'Date': date_str,
                            'Amount': amount,
                            'Description': description,
                            'Category': category,
                            'Subcategory': category.split(' ')[0] if category != 'Other Expenses' else 'Miscellaneous'
                        }
                        all_transactions.append(transaction)
                        sheet_transactions += 1

                    except Exception as e:
                        continue  # Skip problematic rows

                print(f"  [OK] Extracted {sheet_transactions} transactions from {sheet_name}")

            print(f"\n[SUCCESS] Total transactions extracted: {len(all_transactions)}")

            # Apply AI-powered categorization to improve accuracy
            if len(all_transactions) > 0:
                print(f"[AI] Applying AI categorization to {len(all_transactions)} transactions...")
                all_transactions = self.ai_categorize_transactions(all_transactions)
                print(f"[AI] AI categorization complete!")

            return {
                'transactions': all_transactions,
                'bank_info': bank_info,
                'total_rows': len(all_transactions),
                'processing_mode': 'Dynamic Excel Processing with AI Categorization'
            }, None

        except Exception as e:
            print(f"[ERROR] Error processing Excel file: {str(e)}")
            import traceback
            traceback.print_exc()
            return None, f"Error processing Excel file: {str(e)}"

# Global processor instance
processor = UAEBankExcelProcessor()